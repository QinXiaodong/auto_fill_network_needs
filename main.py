from configparser import ConfigParser
import openpyxl
from selenium import webdriver
from selenium.webdriver.support.wait import WebDriverWait  
from selenium.webdriver.support import expected_conditions as EC  
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import re
import time

def getConfig():
    cfg = ConfigParser()
    cfg.read('config.ini', encoding='utf-8')

    username = cfg['DEFAULT']['username']
    password = cfg['DEFAULT']['password']
    rules_file = cfg['DEFAULT']['rules_file']
    return username, password, rules_file

def getRules(rules_file):
    workbook=openpyxl.load_workbook(rules_file) 
    ws=workbook.active
    rules=list()
    row_index = 2
    while(ws.cell(row = row_index, column = 1).value is not None):
        rule = dict()
        for i in range(1, 17):
            rule[ws.cell(row = 1, column = i).value] = ws.cell(row = row_index, column = i).value
        rules.append(rule)
        row_index += 1
    return rules

def autoFill(username, password, rules):
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument('--headless')
    chrome_options.add_argument('--disable-gpu')
    driver = webdriver.Chrome(chrome_options=chrome_options,executable_path='chromedriver.exe')
    driver.set_window_size(1920, 1080)
    driver.get("http://10.245.0.225/uf/login/login.jsp")
    WebDriverWait(driver, 20, 0.5).until(
            EC.presence_of_element_located((By.ID, 'loginbtn')))
    driver.find_element_by_id('username').send_keys(username)
    driver.find_element_by_id('password').send_keys(password)
    driver.find_element_by_id('loginbtn').click()
    WebDriverWait(driver,20,0.5).until(
            EC.presence_of_element_located((By.ID, 'AppFavorite')))
    driver.find_element_by_id('AppFavorite').click()
    WebDriverWait(driver,20,0.5).until(
            EC.presence_of_element_located((By.ID, 'appid_30401')))
    driver.find_element_by_id('appid_30401').click()
    driver.switch_to.window(driver.window_handles[-1])
    driver.set_window_size(1920, 1080)
    WebDriverWait(driver,20,0.5).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, '#app > div > div.sidebar-container.has-logo > div.el-scrollbar > div.scrollbar-wrapper.el-scrollbar__wrap > div > ul > div:nth-child(4) > a')))
    driver.find_element_by_css_selector('#app > div > div.sidebar-container.has-logo > div.el-scrollbar > div.scrollbar-wrapper.el-scrollbar__wrap > div > ul > div:nth-child(4) > a').click()
    WebDriverWait(driver,20,0.5).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, '#app > div > div.main-container > section > div > button')))
    for rule_index in range(len(rules)):
            print('processing rule ' + str(rule_index+1))
            rule = rules[rule_index]
            # 点击新增按钮
            driver.find_element_by_css_selector('#app > div > div.main-container > section > div > button').click()
            WebDriverWait(driver,20,0.5).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, '#app > div > div.main-container > section > div > div.el-dialog__wrapper > div > div.el-dialog__body > div > button:nth-child(2)')))
            # 填充源主机ip地址
            driver.find_element_by_css_selector('#app > div > div.main-container > section > div > div.el-dialog__wrapper > div > div.el-dialog__body > form > div:nth-child(1) > div:nth-child(1) > div > div > div > input').send_keys(rule['源主机ip地址'])

            # 填充目标主机ip地址
            driver.find_element_by_css_selector('#app > div > div.main-container > section > div > div.el-dialog__wrapper > div > div.el-dialog__body > form > div:nth-child(1) > div:nth-child(3) > div > div > div > input').send_keys(rule['目标主机ip地址'])

            # 填充源主机所属系统
            driver.find_element_by_css_selector('#app > div > div.main-container > section > div > div.el-dialog__wrapper > div > div.el-dialog__body > form > div:nth-child(2) > div:nth-child(1) > div > div > div > input').send_keys(rule['源主机所属系统'])

            # 填充目标主机所属系统
            driver.find_element_by_css_selector('#app > div > div.main-container > section > div > div.el-dialog__wrapper > div > div.el-dialog__body > form > div:nth-child(2) > div:nth-child(3) > div > div > div > input').send_keys(rule['目标主机所属系统'])

            # 填充源系统业务负责人
            driver.find_element_by_css_selector('#app > div > div.main-container > section > div > div.el-dialog__wrapper > div > div.el-dialog__body > form > div:nth-child(4) > div:nth-child(1) > div > div > div > input').send_keys(rule['源系统业务负责人'])

            # 填充目标系统业务负责人
            driver.find_element_by_css_selector('#app > div > div.main-container > section > div > div.el-dialog__wrapper > div > div.el-dialog__body > form > div:nth-child(4) > div:nth-child(3) > div > div > div > input').send_keys(rule['目标系统业务负责人'])

            # 填充源负责人联系电话
            driver.find_element_by_css_selector('#app > div > div.main-container > section > div > div.el-dialog__wrapper > div > div.el-dialog__body > form > div:nth-child(5) > div:nth-child(1) > div > div > div > input').send_keys(str(rule['源负责人联系电话']))

            # 填充目标负责人联系电话
            driver.find_element_by_css_selector('#app > div > div.main-container > section > div > div.el-dialog__wrapper > div > div.el-dialog__body > form > div:nth-child(5) > div:nth-child(3) > div > div > div > input').send_keys(str(rule['目标负责人联系电话']))

            # 填充目的端口
            driver.find_element_by_css_selector('#app > div > div.main-container > section > div > div.el-dialog__wrapper > div > div.el-dialog__body > form > div:nth-child(7) > div:nth-child(1) > div > div > div > input').send_keys(rule['目的端口'])

            # 填充宽带需求
            driver.find_element_by_css_selector('#app > div > div.main-container > section > div > div.el-dialog__wrapper > div > div.el-dialog__body > form > div:nth-child(8) > div:nth-child(1) > div > div > div > input').send_keys(rule['宽带需求'])

            # 填充数据交互内容
            driver.find_element_by_css_selector('#app > div > div.main-container > section > div > div.el-dialog__wrapper > div > div.el-dialog__body > form > div:nth-child(8) > div:nth-child(3) > div > div > div > input').send_keys(rule['数据交互内容'])

            # 点选源主机所属机房
            driver.find_element_by_css_selector('#app > div > div.main-container > section > div > div.el-dialog__wrapper > div > div.el-dialog__body > form > div:nth-child(3) > div:nth-child(1) > div > div > div > div.el-input.el-input--suffix > input').click()

            time.sleep(1)
            # 查找总部运营域动态元素并点击
            flag = True
            while(flag):
                    menus = []
                    while(len(menus) <= 1):
                            menus = re.findall('cascader-menu-\d*-0-0', driver.page_source)
                    for menus_index in range(len(menus)):
                            try:
                                    driver.find_element_by_xpath('//*[@id="'+menus[menus_index]+'"]').click()
                                    flag = False
                                    break
                            except:
                                    pass
                    
            # 指定选择项 0：其他， 1：郑州， 2：廊坊
            if(rule['源主机所属机房'] == '郑州'):
                chosen_item = 1  
            elif(rule['源主机所属机房'] == '廊坊'):
                chosen_item = 2
            else:
                chosen_item = 0

            # 查找源主机总部运营域动态子元素并点击
            flag = True
            while(flag):
                    sub_items = []
                    while(len(sub_items) == 0):
                            sub_items=re.findall('cascader-menu-\d*-1-'+str(chosen_item), driver.page_source)
                    for sub_items_index in range(len(sub_items)):
                            try:       
                                    driver.find_element_by_xpath('//*[@id="'+sub_items[sub_items_index]+'"]').click()
                                    flag = False
                                    break
                            except:
                                    pass

            # 点选目标主机所属机房
            driver.find_element_by_css_selector('#app > div > div.main-container > section > div > div.el-dialog__wrapper > div > div.el-dialog__body > form > div:nth-child(3) > div:nth-child(3) > div > div > div > div > input').click()
            time.sleep(1)

            # 点击总部运营域动态元素
            flag = True
            while(flag):
                    for menus_index in range(len(menus)):
                            try:
                                    driver.find_element_by_xpath('//*[@id="'+menus[menus_index]+'"]').click()
                                    flag = False
                                    break
                            except:
                                    pass

            # 指定选择项 0：其他， 1：郑州， 2：廊坊
            if(rule['目标主机所属机房'] == '郑州'):
                    chosen_item = 1  
            elif(rule['目标主机所属机房'] == '廊坊'):
                chosen_item = 2
            else:
                chosen_item = 0
                
            # 查找目标主机总部运营域动态子元素并点击
            flag = True
            while(flag):
                    sub_items = []
                    while(len(sub_items) == 0):
                            sub_items=re.findall('cascader-menu-\d*-1-'+str(chosen_item),driver.page_source)

                    for sub_items_index in range(len(sub_items)):
                            try:       
                                    driver.find_element_by_xpath('//*[@id="'+sub_items[sub_items_index]+'"]').click()
                                    flag = False
                                    break
                            except:
                                    pass
                    
            # 点击源主机所属区域
            driver.find_element_by_css_selector('#app > div > div.main-container > section > div > div.el-dialog__wrapper > div > div.el-dialog__body > form > div:nth-child(6) > div:nth-child(1) > div > div > div > div > input').click()
            time.sleep(1)

            # 选择子区域： 1：核心区， 2:用户访问区， 3：大数据区， 4：DMZ区， 5：测试区， 6：管理维护区， 7：外部系统接口区，8其他
            if(rule['源主机所属区域'] == '核心区'):
                chosen_item = 1
            elif(rule['源主机所属区域'] == '用户访问区'):
                chosen_item = 2
            elif(rule['源主机所属区域'] == '大数据区'):
                chosen_item = 3
            elif(rule['源主机所属区域'] == 'DMZ区'):
                chosen_item = 4
            elif(rule['源主机所属区域'] == '测试区'):
                chosen_item = 5
            elif(rule['源主机所属区域'] == '管理维护区'):
                chosen_item = 6
            elif(rule['源主机所属区域'] == '外部系统接口区'):
                chosen_item = 7
            else:
                chosen_item = 8

            flag = True
            while(flag):
                    for item_index in range(5, 20):
                            try:
                                    driver.find_element_by_xpath('/html/body/div[' + str(item_index) + ']/div[1]/div[1]/ul/li[' + str(chosen_item) + ']') .click()
                                    flag = False
                                    break
                            except:
                                    pass


            # 点击目标主机所属区域
            driver.find_element_by_css_selector('#app > div > div.main-container > section > div > div.el-dialog__wrapper > div > div.el-dialog__body > form > div:nth-child(6) > div:nth-child(3) > div > div > div > div > input').click()

            time.sleep(1)
            # 选择子区域： 1：核心区， 2:用户访问区， 3：大数据区， 4：DMZ区， 5：测试区， 6：管理维护区， 7：外部系统接口区，8其他
            if(rule['目标主机所属区域'] == '核心区'):
                chosen_item = 1
            elif(rule['目标主机所属区域'] == '用户访问区'):
                chosen_item = 2
            elif(rule['目标主机所属区域'] == '大数据区'):
                chosen_item = 3
            elif(rule['目标机所属区域'] == 'DMZ区'):
                chosen_item = 4
            elif(rule['目标机所属区域'] == '测试区'):
                chosen_item = 5
            elif(rule['目标主机所属区域'] == '管理维护区'):
                chosen_item = 6
            elif(rule['目标主机所属区域'] == '外部系统接口区'):
                chosen_item = 7
            else:
                chosen_item = 8

            flag = True
            while(flag):
                    for item_index in range(5, 20):
                            try:
                                    driver.find_element_by_xpath('/html/body/div[' + str(item_index) + ']/div[1]/div[1]/ul/li[' + str(chosen_item) + ']').click()
                                    flag = False
                                    break
                            except:
                                    pass
            # 点击协议
            driver.find_element_by_css_selector('#app > div > div.main-container > section > div > div.el-dialog__wrapper > div > div.el-dialog__body > form > div:nth-child(7) > div:nth-child(3) > div > div > div > div > input').click()

            time.sleep(1)
            # 选择协议 1：tcp, 2: udp
            if(rule['协议'] == 'TCP'):
                chosen_item = 1
            else:
                chosen_item = 2

            flag = True
            while(flag):
                    for item_index in range(5, 20):
                            try:
                                    driver.find_element_by_xpath('/html/body/div[' + str(item_index)+ ']/div[1]/div[1]/ul/li[' + str(chosen_item) + ']').click()
                                    flag = False
                                    break
                            except:
                                    pass

            driver.save_screenshot('screen_shot2.png')
            #点击保存

            driver.find_element_by_css_selector('#app > div > div.main-container > section > div > div.el-dialog__wrapper > div > div.el-dialog__body > div > button:nth-child(2)').click()

    # 滚动到页面到底部
    js="var q=document.documentElement.scrollTop=100000"
    driver.execute_script(js)

    # 点击保存草稿
    driver.find_element_by_xpath('//*[@id="app"]/div/div[2]/section/div/div[2]/button[2]').click()
    time.sleep(2)
    driver.quit()

def main():
    wo_username, wo_password, rules_file = getConfig()
    # print(wo_username)
    # print(wo_password)
    rules = getRules(rules_file)
    autoFill(wo_username, wo_password, rules)
    print('done!')

if __name__ == '__main__':
    main()